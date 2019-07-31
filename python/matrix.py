# https://en.wikipedia.org/wiki/Digital_differential_analyzer_(graphics_algorithm)
try:
    import openpyxl
    from openpyxl.worksheet.write_only import WriteOnlyCell
    from openpyxl.styles import PatternFill
except ModuleNotFoundError:
    openpyxl = None
try:
    import pygame
    from pygame.locals import *
except ModuleNotFoundError:
    pygame = None

import math

# x = col
# y = row
class Matrix:

    def __init__(self, width, height=None, default=0):
        self.width = width
        self.height = height if height is not None else width
        self.data = []
        for x in range(0, self.width):
            self.data.append([])
            for y in range(0, self.height):
                self.data[-1].append(default)

    def set(self, x, y, v):
        self.data[x][y] = v

    def set_rect(self, x, y, w, h, v):
        for col in range(x, x + w):
            for row in range(y, y + h):
                self.data[col][row] = v
    
    def get(self, x, y):
        return self.data[x][y]
    
    def fill(self, v):
        for y in range(0, self.height):
            for x in range(0, self.width):
                self.data[x][y] = v

    def print(self):
        self.to_console()
    
    def to_console(self):
        for y in range(0, self.height):
            for x in range(0, self.width):
                print(self.data[x][y], '', end='')
            print()
    
    def to_excel(self, title="output.xlsx"):
        if openpyxl is None:
            raise Exception("Module openpyxl not available.")
        else:
            wb = openpyxl.Workbook(write_only=True)
            ws = wb.create_sheet("One")
            ws.sheet_format.defaultColWidth = 3.0
            ws.defaultRowHeight = 4.0
            fill = PatternFill("solid", fgColor="FF0000")
            for y in range(0, self.height):
                row = []
                for x in range(0, self.width):
                    cell = WriteOnlyCell(ws, value=self.data[x][y])
                    cell.fill = fill
                    row.append(cell)
                    #row.append(self.data[x][y])
                ws.append(row)
            wb.save(title)

    def to_image(self, filename="output.png", colors=None, factor=32):
        if pygame is None:
            raise Exception("Module pygame not available.")
        else:
            #pygame.init()
            pygame.font.init()
            font = pygame.font.SysFont("arial", 16)
            surf = pygame.Surface((factor * self.width, factor * self.height))
            for y in range(0, self.height):
                for x in range(0, self.width):
                    color = (255, 0, 0) if colors is None else colors[self.data[x][y]]
                    pygame.draw.rect(surf, color, (x*factor, y*factor, factor, factor), 0)
                    pygame.draw.rect(surf, (255, 255, 255), (x*factor, y*factor, factor, factor), 1)
                    tex = font.render(str(self.data[x][y]), True, (255, 255, 255))
                    surf.blit(tex, (x*factor+10, y*factor+8))
            if filename is not None: pygame.image.save(surf, filename)
            return surf
    
    def to_screen(self, screen):
        font = pygame.font.SysFont("arial", 16)
        colors = {0 : (0, 0, 0), 1 : (255, 0, 0), 2 : (0, 255, 0), 3 : (0, 0, 255), 4 : (255, 255, 0)}
        factor = 40
        for y in range(0, self.height):
            for x in range(0, self.width):
                color = colors[self.data[x][y]]
                pygame.draw.rect(screen, color, (x*factor, y*factor, factor, factor), 0)
                pygame.draw.rect(screen, (255, 255, 255), (x*factor, y*factor, factor, factor), 1)
                tex = font.render(str(self.data[x][y]), True, (255, 255, 255))
                screen.blit(tex, (x*factor+10, y*factor+8))


def line(start, end, filename):
    matrix = Matrix(10)
    surf = matrix.to_image(
        filename = None,
        colors = {0 : (0, 0, 0), 1 : (255, 0, 0), 2 : (0, 255, 0), 3 : (0, 0, 255), 4 : (255, 255, 0)},
        factor = 40
    )
    # real coord
    x, y = start[0], start[1]
    end_x, end_y = end[0], end[1]
    # map coord
    map_x = int(x)
    map_y = int(y)
    end_map_x = int(end_x)
    end_map_y = int(end_y)
    matrix.set(map_x, map_y, 2)
    matrix.set(end_map_x, end_map_y, 1)
    diff_x = end_x - x
    diff_y = end_y - y
    if abs(diff_x) > abs(diff_y):
        delta_x = 1 if diff_x > 0 else -1
        delta_y = diff_y / abs(diff_x)
    else:
        delta_x = diff_x / abs(diff_y)
        delta_y = 1 if diff_y > 0 else -1
    pygame.draw.rect(surf, (120, 120, 120), (map_x * 40 + 1, map_y * 40 + 1, 40 - 2, 40 - 2), 0)
    pygame.draw.circle(surf, (0, 0, 255), (int(x * 40), int(y * 40)), 5, 0)
    while map_x != end_map_x or map_y != end_map_y:
        print('------------------')
        print(f'     x = {round(x, 1):3.1f}      y = {round(y, 1):3.1f}')
        x += delta_x
        y += delta_y
        map_x = int(x)
        map_y = int(y)
        print('map x =', map_x)
        print('map y =', map_y)
        pygame.draw.rect(surf, (120, 120, 120), (map_x * 40 + 1, map_y * 40 + 1, 40 - 2, 40 - 2), 0)
        pygame.draw.circle(surf, (255, 0, 0), (int(x * 40), int(y * 40)), 5, 0)
        matrix.set(map_x, map_y, 3)
    print('------------------')
    print(f'     x = {round(x, 1):3.1f}      y = {round(y, 1):3.1f}')
    pygame.draw.circle(surf, (0, 0, 255), (int(end[0] * 40), int(end[1] * 40)), 5, 0)
    pygame.draw.line(surf, (255, 255, 0), (start[0] * 40, start[1] * 40), (end[0] * 40, end[1] * 40))
    pygame.image.save(surf, filename)


def line_calc(matrix, x, y, end_x, end_y):
    map_x = int(x)
    map_y = int(y)
    diff_x = end_x - x
    diff_y = end_y - y
    if abs(diff_x) > abs(diff_y):
        delta_x = 1 if diff_x > 0 else -1
        delta_y = diff_y / abs(diff_x)
    else:
        delta_x = diff_x / abs(diff_y)
        delta_y = 1 if diff_y > 0 else -1
    while matrix.get(map_x, map_y) != 1:
        x += delta_x
        y += delta_y
        old_map_x = map_x
        old_map_y = map_y
        map_x = int(x)
        map_y = int(y)
        print(map_x, map_y, matrix.get(map_x, map_y))
    print('-------------')
    return None


ORANGE      = (255, 165,   0)
WHITE       = (255, 255, 255)
RED         = (255,   0,   0)
DARK_ORANGE = (255,  69,   0)
BLUE        = (  0,   0, 255)
GREEN       = (  0, 255,   0)
YELLOW      = (255, 255,   0)
BLACK       = (  0,   0,   0)

def dist(x1, y1, x2, y2):
    return math.sqrt((x2-x1)**2 + (y2-y1)**2)



def line_draw(surf, matrix, x, y, end_x, end_y, a):
    map_x = int(x)
    map_y = int(y)
    v_x = end_x - x
    v_y = end_y - y
    next_map_x = map_x
    step_x = 0
    step_y = 0
    if v_x > 0:
        next_map_x = map_x + 1
        step_x = 1
    elif v_x < 0:
        next_map_x = map_x
        step_x = -1
    next_map_y = map_y
    if v_y > 0:
        next_map_y = map_y + 1
        step_y = 1
    elif v_y < 0:
        next_map_y = map_y
        step_y = -1
    cpt = 1
    next_point = None
    while True:
        if abs(end_x - x) > 0.0000000000001 and abs(end_y - y) > 0.0000000000001:
            a = (end_y - y) / (end_x - x)
            b = y - a * x
            next_hori = ((next_map_y - b) / a, next_map_y)
            next_vert = (next_map_x, a * next_map_x + b)
            if dist(x, y, next_hori[0], next_hori[1]) < dist(x, y, next_vert[0], next_hori[1]):
                next_point = next_hori
            else:
                next_point = next_vert
        elif abs(end_x - x) <= 0.0000000000001 :
            next_hori = (x, next_map_y)
            next_vert = None
            next_point = next_hori
        elif abs(end_y - y) <= 0.0000000000001:
            next_hori = None
            next_vert = (next_map_x, y)
            next_point = next_vert
        if next_hori == next_vert:
            next_map_x += step_x
            next_map_y += step_y
        elif next_point == next_hori:
            next_map_y += step_y
        else:
            next_map_x += step_x
        if next_point == next_hori:
            if step_y > 0:
                if matrix.get(int(next_point[0]), int(next_point[1])) == 1:
                    break
            else:
                if matrix.get(int(next_point[0]), int(next_point[1] - 1)) == 1:
                    break
        else:
            if step_x > 0:
                if matrix.get(int(next_point[0]), int(next_point[1])) == 1:
                    break
            else:
                if matrix.get(int(next_point[0]) - 1, int(next_point[1])):
                    break
    pygame.draw.line(surf, YELLOW, (tos(x), tos(y)), tos2(next_point))
    d = dist(x, y, next_point[0], next_point[1])
    if next_point == next_hori:
        d *= math.cos(a)
    else:
        d *= math.cos(a)
    return next_point[0], next_point[1], d


def line_draw_full(surf, matrix, x, y, end_x, end_y, a):
    pygame.draw.circle(surf, (0, 255, 0), (tos(x), tos(y)), 5, 0)
    map_x = int(x)
    map_y = int(y)
    v_x = end_x - x
    v_y = end_y - y
    next_map_x = map_x
    step_x = 0
    step_y = 0
    if v_x > 0:
        next_map_x = map_x + 1
        step_x = 1
    elif v_x < 0:
        next_map_x = map_x
        step_x = -1
    next_map_y = map_y
    if v_y > 0:
        next_map_y = map_y + 1
        step_y = 1
    elif v_y < 0:
        next_map_y = map_y
        step_y = -1
    cpt = 1
    next_point = None
    while True:
        if abs(end_x - x) > 0.0000000000001 and abs(end_y - y) > 0.0000000000001:
            a = (end_y - y) / (end_x - x)
            b = y - a * x
            #obj = 100 if v_x > 0 else -100
            #horizon = (obj, a * obj + b)
            next_hori = ((next_map_y - b) / a, next_map_y)
            next_vert = (next_map_x, a * next_map_x + b)
            #diff_hori_x = abs(next_hori[0] - x)
            #diff_hori_y = 1
            #diff_vert_x = 1
            #diff_vert_y = abs(next_vert[0] - y)
            # On doit choisir entre next_hori et next_vert
            if dist(x, y, next_hori[0], next_hori[1]) < dist(x, y, next_vert[0], next_hori[1]):
                next_point = next_hori
            else:
                next_point = next_vert
        elif abs(end_x - x) <= 0.0000000000001 :
            #obj = 100 if v_y > 0 else -100
            #horizon = (x, obj)
            next_hori = (x, next_map_y)
            next_vert = None
            next_point = next_hori
        elif abs(end_y - y) <= 0.0000000000001:
            #obj = 100 if v_x > 0 else -100
            #horizon = (obj, y)
            next_hori = None
            next_vert = (next_map_x, y)
            next_point = next_vert
        if next_hori == next_vert:
            next_map_x += step_x
            next_map_y += step_y
        elif next_point == next_hori:
            next_map_y += step_y
        else:
            next_map_x += step_x
        nx = " hori" if next_point == next_hori else " vert"
        tex = font.render(str(cpt) + nx, False, WHITE)
        cpt+=1
        surf.blit(tex, (tos2(next_point)[0], tos2(next_point)[1])) # + 20
        pygame.draw.circle(surf, GREEN, tos2(next_point), 5, 0)
        if next_point == next_hori:
            if step_y > 0:
                if matrix.get(int(next_point[0]), int(next_point[1])) == 1:
                    break
            else:
                if matrix.get(int(next_point[0]), int(next_point[1] - 1)) == 1:
                    break
        else:
            if step_x > 0:
                if matrix.get(int(next_point[0]), int(next_point[1])) == 1:
                    break
            else:
                if matrix.get(int(next_point[0]) - 1, int(next_point[1])):
                    break
        # Draw
        #pygame.draw.line(surf, (255, 255, 0), (tos(x), tos(y)), tos2(horizon))
        #if next_vert is not None:
        #    tex = font.render("vert", False, ORANGE)
        #    surf.blit(tex, tos2(next_vert))
        #    pygame.draw.circle(surf, (0, 255, 0), tos2(next_vert), 5, 0)
        #if next_hori is not None:
        #    tex = font.render("hori", False, ORANGE)
        #    surf.blit(tex, tos2(next_hori))
        #    pygame.draw.circle(surf, (0, 0, 255), tos2(next_hori), 5, 0)
        #pygame.draw.circle(surf, (0, 255, 0), (tos(next_map_x) + 20, tos(next_map_y) + 20), 5, 0)
    pygame.draw.line(surf, YELLOW, (tos(x), tos(y)), tos2(next_point))
    d = dist(x, y, next_point[0], next_point[1])
    if next_point == next_hori:
        d *= math.cos(a)
    else:
        d *= math.cos(a)
    return next_point[0], next_point[1], d


def line_draw1(surf, matrix, x, y, end_x, end_y):
    pygame.draw.circle(surf, (0, 255, 0), (tos(x), tos(y)), 5, 0)
    map_x = int(x)
    map_y = int(y)
    diff_x = end_x - x
    diff_y = end_y - y
    if abs(diff_x) > abs(diff_y):
        delta_x = 1 if diff_x > 0 else -1
        delta_y = diff_y / abs(diff_x)
    else:
        delta_x = diff_x / abs(diff_y)
        delta_y = 1 if diff_y > 0 else -1
    while matrix.get(map_x, map_y) != 1:
        x += delta_x
        y += delta_y
        old_map_x = map_x
        old_map_y = map_y
        map_x = int(x)
        map_y = int(y)
        pygame.draw.circle(surf, (0, 255, 0), (tos(x), tos(y)), 5, 0)
    return None


def tos2(v):
    return (tos(v[0]), tos(v[1]))


def tos(v):
    return int(v*40)

font = None
debug = False

def main():
    global font, debug
    SCREEN_WIDTH = 200
    SCREEN_HEIGHT = 150
    player_x = 3.5
    player_y = 2.5
    #player_a = math.pi / 2
    player_a = 0
    fov = 66 * math.pi / 180
    step_a = fov / SCREEN_WIDTH
    map_x = int(player_x)
    map_y = int(player_y)
    speed_move = 0.01
    speed_turn = 0.01
    pygame.init()
    screen = pygame.display.set_mode((640, 480), 0, 32)
    font = pygame.font.SysFont("arial", 16)
    m = Matrix(10)
    m.set_rect(0, 0, 10, 1, 1)
    m.set_rect(0, 0, 1, 10, 1)
    m.set_rect(0, 9, 10, 1, 1)
    m.set_rect(9, 0, 1, 10, 1)
    m.set(6, 4, 1)
    m.set(6, 5, 1)
    m.set(6, 6, 1)
    columns = [0] * SCREEN_WIDTH
    while True:
        # -- Input --
        for event in pygame.event.get():
            if event.type == QUIT:
                pygame.quit()
                return
            elif event.type == KEYUP:
                if event.key == K_SPACE:
                    debug = not debug
        keys = pygame.key.get_pressed()
        m.set(int(player_x), int(player_y), 0)
        old_x = player_x
        old_y = player_y
        old_map_x = map_x
        old_map_y = map_y
        if keys[K_LEFT]:
            player_a -= speed_turn
            if player_a < 0: player_a += 2 * math.pi
        if keys[K_RIGHT]:
            player_a += speed_turn
            if player_a >= 2 * math.pi: player_a -= 2 * math.pi
        if keys[K_UP]:
            player_x += math.cos(player_a) * speed_move
            player_y += math.sin(player_a) * speed_move
        if keys[K_DOWN]:
            player_x -= math.cos(player_a) * speed_move
            player_y -= math.sin(player_a) * speed_move
        map_x = int(player_x)
        map_y = int(player_y)
        if m.get(map_x, map_y) != 0:
            player_x = old_x
            player_y = old_y
            map_x = old_map_x
            map_y = old_map_y
        m.set(map_x, map_y, 4)
        # -- Update --
        a = player_a
        #ray_x = int(player_x + math.cos(a))
        #ray_y = int(player_y + math.sin(a))
        #lineX(m, player_x, player_y, ray_x, ray_y)
        
        #a = player_a
        #for i in range(0, 1):
        a = player_a - fov / 2
        for i in range(0, SCREEN_WIDTH):
            ray_x = int(player_x + math.cos(a))
            ray_y = int(player_y + math.sin(a))
            #lineX(m, (ray_x, ray_y))
            #columns[i] = SCREEN_HEIGHT
            columns[SCREEN_WIDTH // 2] = SCREEN_HEIGHT
            a += step_a
        # -- Draw --
        screen.fill(BLACK)
        m.to_screen(screen)
        # 2D screen
        screen_x = int(player_x * 40)
        screen_y = int(player_y * 40)
        ray_x = int((player_x + math.cos(player_a)) * 40)
        ray_y = int((player_y + math.sin(player_a)) * 40)
        screen.set_at((screen_x, screen_y), (0, 0, 255))
        pygame.draw.line(screen, (0, 0, 255), (screen_x, screen_y), (ray_x, ray_y))
        if debug:
            a = player_a
            for i in range(0, 1):
                ray_x = player_x + math.cos(a)
                ray_y = player_y + math.sin(a)
                # columns[i] = SCREEN_HEIGHT
                a += step_a
        else:
            a = player_a - fov / 2
            for i in range(0, SCREEN_WIDTH):
                ray_x = player_x + math.cos(a)
                ray_y = player_y + math.sin(a)
                columns[i] = SCREEN_HEIGHT / d
                a += step_a
        # 2.5D screen
        #pygame.draw.rect(screen, (0, 255, 0), (410, 10, SCREEN_WIDTH, SCREEN_HEIGHT), 0)
        for i in range(0, SCREEN_WIDTH):
            deb_line_y = SCREEN_HEIGHT / 2 - columns[i] / 2
            if deb_line_y < 0: deb_line_y = 0
            if deb_line_y > SCREEN_HEIGHT / 2: deb_line_y = 0
            end_line_y = SCREEN_HEIGHT / 2 + columns[i] / 2
            if end_line_y > SCREEN_HEIGHT: end_line_y = SCREEN_HEIGHT
            if end_line_y < SCREEN_HEIGHT / 2: end_line_y = SCREEN_HEIGHT
            try:
                pygame.draw.line(screen, GREEN, (i + 410, deb_line_y), (i + 410, end_line_y))
            except TypeError:
                print(deb_line_y)
                print(end_line_y)
                raise Exception("pipo")
        # Info
        tex = font.render("a = " + str(round(player_a, 1)), False, (255, 255, 0))
        screen.blit(tex, (410, 220))
        tex = font.render("x = " + str(round(player_x, 1)), False, (255, 255, 0))
        screen.blit(tex, (410, 240))
        tex = font.render("y = " + str(round(player_y, 1)), False, (255, 255, 0))
        screen.blit(tex, (410, 260))
        pygame.display.update()


if __name__ == '__main__':
    #m = Matrix(10)
    #m.set(0, 0, 1)
    #m.set(9, 0, 2)
    #m.set(9, 9, 3)
    #m.set(0, 9, 4)
    #m.print()
    #m.to_excel()
    #surf = m.to_image(
    #    filename = "zorba.png",
    #    colors = {0 : (0, 0, 0), 1 : (255, 0, 0), 2 : (0, 255, 0), 3 : (0, 0, 255), 4 : (255, 255, 0)},
    #    factor = 40
    #)
    #line(surf, m, (1.3, 2.2), (7.4, 5.7))
    #line((1.3, 2.2), (6.4, 4.7), "t1.png")
    #line((6.4, 4.7), (1.3, 2.2), "t2.png")
    #line((1.7, 4.7), (6.7, 4.7), "t3.png")
    #m.fill(5)
    #m.print()
    main()
